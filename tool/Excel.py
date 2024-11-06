from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


class Excel:
    border = Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))

    def __init__(self):
        self.excel = None
        self.sheet = None
        self.sheet2 = None

    def work_book(self):
        wb = Workbook()
        wb.remove(wb['Sheet'])
        self.sheet = wb.create_sheet("结果", 0)
        self.sheet2 = wb.create_sheet("可疑物种", 1)
        self.excel = wb
        return self

    def write_line(self, data: list[str]):
        self.sheet.append(data)
        return self

    def set_title_border(self, m: list[str]):
        self.sheet.append(m)
        self.sheet2.append(m)
        for i in range(1, len(m) + 1):
            self.sheet[get_column_letter(i) + "1"].border = self.border
            self.sheet2[get_column_letter(i) + "1"].border = self.border
        return self

    def save(self, o):
        self.excel.save(o)

